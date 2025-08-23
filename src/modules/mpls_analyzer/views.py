from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django_ratelimit.decorators import ratelimit
 
# MFA temporarily disabled for development
# from django_otp import user_has_device
# from django_otp.plugins.otp_totp.models import TOTPDevice
 
from django.utils import timezone
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import pyotp
import qrcode
from io import BytesIO
import base64
 
from base64 import b32encode

from .models import (
    Equipment, CustomerService, Vpn, 
    BackupProcessLog, Interface, UserProfile,
    AccessLog, AuditLog, SecuritySettings, LoginAttempt
)
 
from django.views.decorators.http import require_GET
from .security import (
    require_mfa, track_login_attempts, clear_login_attempts,
    log_security_event, get_client_ip, setup_mfa_for_user,
    generate_qr_code, validate_password_strength, is_manager
)
from .forms import (
    UserRegistrationForm, UserProfileForm, StrongPasswordChangeForm,
    AdvancedSearchForm
)
from .audit import AuditLogger, log_audit_action
from .search_utils import smart_search, AdvancedSearchEngine
@require_GET
@require_mfa
def customer_interface_report(request):
    """Relatório de interfaces de clientes por equipamento"""
    equipment_name = request.GET.get('equipment')
    
    interfaces_query = Interface.objects.filter(is_customer_interface=True).select_related(
        'mpls_config__equipment'
    ).prefetch_related('members')
    
    if equipment_name:
        interfaces_query = interfaces_query.filter(
            mpls_config__equipment__name__icontains=equipment_name
        )
    
    results = []
    for interface in interfaces_query:
        equipment = interface.mpls_config.equipment
        
        # Busca VPNs associadas a esta interface
        vpns = Vpn.objects.filter(
            vpws_group__mpls_config=interface.mpls_config,
            access_interface=interface.name
        ).select_related('vpws_group')
        
        interface_data = {
            'equipment': {
                'hostname': equipment.name,
                'loopback_ip': equipment.ip_address,
                'location': equipment.location
            },
            'interface': {
                'name': interface.name,
                'description': interface.description,
                'type': interface.interface_type,
                'speed': interface.speed
            },
            'vpns': [],
            'customers': []
        }
        
        # LAG members se aplicável
        if interface.interface_type == 'lag':
            interface_data['interface']['lag_members'] = list(
                interface.members.values_list('member_interface_name', flat=True)
            )
        
        # Dados das VPNs
        for vpn in vpns:
            vpn_data = {
                'vpn_id': vpn.vpn_id,
                'neighbor_ip': vpn.neighbor_ip,
                'neighbor_hostname': _resolve_equipment_by_loopback(vpn.neighbor_ip),
                'description': vpn.description,
                'encapsulation': vpn.encapsulation,
                'encapsulation_type': vpn.encapsulation_type,
                'customers': [cs.name for cs in vpn.customer_services.all()]
            }
            interface_data['vpns'].append(vpn_data)
            interface_data['customers'].extend(vpn_data['customers'])
        
        # Remove duplicatas de clientes
        interface_data['customers'] = list(set(interface_data['customers']))
        results.append(interface_data)
    
    # Registra auditoria da consulta de interfaces de clientes
    log_audit_action(
        user=request.user,
        action='report_export',
        description=f'Relatório de interfaces de clientes consultado{f" - equipamento: {equipment_name}" if equipment_name else ""}',
        request=request,
        search_query=equipment_name or '',
        results_count=len(results),
        additional_data={
            'report_type': 'customer_interface_report',
            'equipment_name': equipment_name or 'all'
        }
    )
    
    return JsonResponse({'results': results})


@require_GET
def vpn_report(request):
    """Relatório por VPN ID: ponta A e ponta B, interface de acesso e descrições"""
    vpn_id = request.GET.get('vpn_id')
    if not vpn_id:
        return JsonResponse({'error': 'vpn_id obrigatório'}, status=400)
    try:
        vpn_id_int = int(vpn_id)
    except ValueError:
        return JsonResponse({'error': 'vpn_id inválido'}, status=400)

    # Busca VPNs com este ID (pode haver múltiplas instâncias em diferentes grupos/configs)
    vpns = (
        Vpn.objects.filter(vpn_id=vpn_id_int)
        .select_related('vpws_group__mpls_config__equipment')
    )

    results = []
    for v in vpns:
        equip_a = v.vpws_group.mpls_config.equipment
        
        # Busca detalhes da interface de acesso
        access_interface_details = None
        if v.access_interface:
            try:
                interface = Interface.objects.get(
                    mpls_config=v.vpws_group.mpls_config,
                    name=v.access_interface
                )
                access_interface_details = {
                    'name': interface.name,
                    'description': interface.description,
                    'type': interface.interface_type,
                    'speed': interface.speed,
                    'is_customer': interface.is_customer_interface
                }
                
                # Se é LAG, busca membros
                if interface.interface_type == 'lag':
                    members = list(interface.members.values_list('member_interface_name', flat=True))
                    access_interface_details['lag_members'] = members
                    
            except Interface.DoesNotExist:
                pass
        
        # Busca todas as VPNs com o mesmo ID para mostrar correlação ponta-a-ponta
        peer_vpns = Vpn.objects.filter(
            vpn_id=v.vpn_id,
            neighbor_ip=equip_a.ip_address
        ).exclude(id=v.id).select_related('vpws_group__mpls_config__equipment')
        
        peer_equipment = []
        for peer_vpn in peer_vpns:
            peer_equip = peer_vpn.vpws_group.mpls_config.equipment
            peer_equipment.append({
                'hostname': peer_equip.name,
                'loopback_ip': peer_equip.ip_address,
                'access_interface': peer_vpn.access_interface,
                'encapsulation': peer_vpn.encapsulation,
                'encapsulation_type': peer_vpn.encapsulation_type
            })
        
        results.append({
            'vpn_id': v.vpn_id,
            'encapsulation': v.encapsulation,
            'encapsulation_type': v.encapsulation_type,
            'access_interface': v.access_interface,
            'access_interface_details': access_interface_details,
            'description': v.description,
            'pw_type': v.pw_type,
            'pw_id': v.pw_id,
            'equipment_a': {
                'hostname': equip_a.name,
                'loopback_ip': equip_a.ip_address,
                'location': equip_a.location,
            },
            'equipment_b': {
                'loopback_ip': v.neighbor_ip,
                'hostname': _resolve_equipment_by_loopback(v.neighbor_ip),
            },
            'peer_equipment': peer_equipment,
            'customers': [cs.name for cs in v.customer_services.all()]
        })

    # Registra auditoria da consulta de VPN
    log_audit_action(
        user=request.user,
        action='view_vpn',
        description=f'Consultou relatório de VPN: {vpn_id}',
        request=request,
        search_query=str(vpn_id),
        target_object_type='Vpn',
        target_object_id=vpn_id_int,
        results_count=len(results),
        additional_data={
            'report_type': 'vpn_report',
            'vpn_id': vpn_id_int
        }
    )

    return JsonResponse({'results': results})


# =============================================================================
# CONFIGURAÇÕES DE SEGURANÇA
# =============================================================================

@require_mfa
@user_passes_test(is_manager)
def security_settings_view(request):
    """View para configurar as configurações globais de segurança"""
    settings = SecuritySettings.get_settings()
    
    if request.method == 'POST':
        try:
            # Atualiza as configurações com os dados do POST
            settings.max_login_attempts = int(request.POST.get('max_login_attempts', 5))
            settings.lockout_duration_minutes = int(request.POST.get('lockout_duration_minutes', 15))
            settings.session_timeout_minutes = int(request.POST.get('session_timeout_minutes', 120))
            settings.audit_retention_days = int(request.POST.get('audit_retention_days', 90))
            
            # Configurações de IP
            settings.enable_ip_whitelist = request.POST.get('enable_ip_whitelist') == 'on'
            settings.allowed_ips = request.POST.get('allowed_ips', '')
            
            # Configurações de senha
            settings.password_min_length = int(request.POST.get('password_min_length', 8))
            settings.password_require_uppercase = request.POST.get('password_require_uppercase') == 'on'
            settings.password_require_lowercase = request.POST.get('password_require_lowercase') == 'on'
            settings.password_require_numbers = request.POST.get('password_require_numbers') == 'on'
            settings.password_require_symbols = request.POST.get('password_require_symbols') == 'on'
            
            # Metadados
            settings.updated_by = request.user
            settings.save()
            
            # Registra auditoria da alteração
            log_audit_action(
                user=request.user,
                action='system_settings',
                description='Configurações de segurança atualizadas',
                request=request,
                additional_data={
                    'settings_changed': {
                        'max_login_attempts': settings.max_login_attempts,
                        'lockout_duration_minutes': settings.lockout_duration_minutes,
                        'session_timeout_minutes': settings.session_timeout_minutes,
                    }
                }
            )
            
            messages.success(request, 'Configurações de segurança atualizadas com sucesso!')
            return redirect('security_settings')
            
        except (ValueError, TypeError) as e:
            messages.error(request, f'Erro nos dados fornecidos: {str(e)}')
        except Exception as e:
            messages.error(request, f'Erro interno: {str(e)}')
    
    # Estatísticas de segurança
    from datetime import timedelta
    from django.utils import timezone
    
    last_24h = timezone.now() - timedelta(hours=24)
    last_7d = timezone.now() - timedelta(days=7)
    
    security_stats = {
        'total_login_attempts_24h': LoginAttempt.objects.filter(timestamp__gte=last_24h).count(),
        'failed_attempts_24h': LoginAttempt.objects.filter(timestamp__gte=last_24h, success=False).count(),
        'blocked_ips_24h': LoginAttempt.objects.filter(
            timestamp__gte=last_24h,
            failure_reason='Login bloqueado por muitas tentativas'
        ).values('ip_address').distinct().count(),
        'total_attempts_7d': LoginAttempt.objects.filter(timestamp__gte=last_7d).count(),
        'unique_users_7d': LoginAttempt.objects.filter(timestamp__gte=last_7d).values('username').distinct().count(),
    }
    
    context = {
        'settings': settings,
        'security_stats': security_stats,
    }
    
    return render(request, 'mpls_analyzer/security_settings.html', context)


def _resolve_equipment_by_loopback(loopback_ip: str):
    try:
        equip = Equipment.objects.get(ip_address=loopback_ip)
        return equip.name
    except Equipment.DoesNotExist:
        return ''


@ratelimit(key='ip', rate='5/m', method='POST', block=True)
def login_view(request):
    if request.user.is_authenticated:
        # MFA disabled for development - go directly to dashboard
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        ip_address = get_client_ip(request)
        user_agent = request.META.get('HTTP_USER_AGENT', '')
        
        if username and password:
            # Verifica se o usuário ou IP está bloqueado
            is_ip_blocked = LoginAttempt.is_ip_blocked(ip_address)
            is_user_blocked = LoginAttempt.is_user_blocked(username)
            
            if is_ip_blocked or is_user_blocked:
                settings = SecuritySettings.get_settings()
                
                # Registra tentativa bloqueada
                LoginAttempt.record_attempt(
                    username=username,
                    ip_address=ip_address,
                    user_agent=user_agent,
                    success=False,
                    failure_reason='Login bloqueado por muitas tentativas'
                )
                
                messages.error(
                    request, 
                    f'Login bloqueado por muitas tentativas falhadas. '
                    f'Tente novamente em {settings.lockout_duration_minutes} minutos.'
                )
                log_security_event('LOGIN_BLOCKED', ip_address=ip_address, details=f"Username: {username}")
                return render(request, 'mpls_analyzer/login.html')
            
            # Tenta autenticar o usuário
            user = authenticate(request, username=username, password=password)
            if user is not None:
                # Login bem-sucedido
                LoginAttempt.record_attempt(
                    username=username,
                    ip_address=ip_address,
                    user_agent=user_agent,
                    success=True
                )
                
                login(request, user)
                log_security_event('LOGIN_SUCCESS', user=user, ip_address=ip_address)
                
                # MFA disabled for development - go directly to dashboard
                return redirect('dashboard')
            else:
                # Login falhou
                LoginAttempt.record_attempt(
                    username=username,
                    ip_address=ip_address,
                    user_agent=user_agent,
                    success=False,
                    failure_reason='Credenciais inválidas'
                )
                
                # Verifica quantas tentativas restam
                settings = SecuritySettings.get_settings()
                from django.utils import timezone
                from datetime import timedelta
                
                check_period = timezone.now() - timedelta(minutes=settings.lockout_duration_minutes)
                recent_failures = LoginAttempt.objects.filter(
                    username=username,
                    success=False,
                    timestamp__gte=check_period
                ).count()
                
                remaining_attempts = settings.max_login_attempts - recent_failures
                
                if remaining_attempts <= 0:
                    messages.error(
                        request, 
                        f'Muitas tentativas falhadas. Login bloqueado por {settings.lockout_duration_minutes} minutos.'
                    )
                else:
                    messages.error(
                        request, 
                        f'Usuário ou senha inválidos. {remaining_attempts} tentativa(s) restante(s).'
                    )
                
                log_security_event('LOGIN_FAILED', ip_address=ip_address, details=f"Username: {username}")
        else:
            messages.error(request, 'Preencha todos os campos.')
    
    return render(request, 'mpls_analyzer/login.html')


def logout_view(request):
    if request.user.is_authenticated:
        log_security_event('LOGOUT', user=request.user, ip_address=get_client_ip(request))
    logout(request)
    return redirect('login')


# MFA views disabled for development
@login_required
def setup_mfa(request):
    """MFA disabled for development - redirect to dashboard"""
    return redirect('dashboard')


@login_required
def verify_mfa(request):
    """MFA disabled for development - redirect to dashboard"""
    return redirect('dashboard')


@require_mfa
def dashboard(request):
    # Estatísticas gerais
    total_equipment = Equipment.objects.count()
    total_customers = CustomerService.objects.values('name').distinct().count()
    
    # Conta VPNs únicas (considerando A-B como uma unidade)
    # Agrupa por VPN ID e conta apenas uma vez
    unique_vpn_ids = Vpn.objects.values_list('vpn_id', flat=True).distinct().count()
    total_vpns = unique_vpn_ids
    
    last_backup = Equipment.objects.filter(last_backup__isnull=False).order_by('-last_backup').first()
    
    # Logs recentes
    recent_logs = BackupProcessLog.objects.all()[:5]
    
    context = {
        'total_equipment': total_equipment,
        'total_customers': total_customers,
        'total_vpns': total_vpns,
        'last_backup': last_backup,
        'recent_logs': recent_logs,
    }
    
    return render(request, 'mpls_analyzer/dashboard.html', context)


@require_mfa
def customer_report_view(request):
    """View para renderizar o template do relatório de cliente"""
    customer_name = request.GET.get('customer', '')
    
    # Registra auditoria do acesso à página de relatório de cliente
    if customer_name:
        log_audit_action(
            user=request.user,
            action='view_equipment',
            description=f'Acessou página de relatório de cliente: {customer_name}',
            request=request,
            search_query=customer_name,
            additional_data={
                'page_type': 'customer_report_view',
                'customer_name': customer_name
            }
        )
    else:
        log_audit_action(
            user=request.user,
            action='view_equipment',
            description='Acessou página de relatório de cliente (sem cliente específico)',
            request=request,
            additional_data={
                'page_type': 'customer_report_view'
            }
        )
    
    context = {
        'customer_name': customer_name,
    }
    
    return render(request, 'mpls_analyzer/customer_report.html', context)


@require_mfa
def customer_report(request):
    """Relatório detalhado de cliente mostrando lado A e B lado a lado"""
    customer_name = request.GET.get('customer')
    if not customer_name:
        return JsonResponse({'error': 'Nome do cliente é obrigatório'}, status=400)
    
    try:
        # Busca todos os serviços do cliente
        customer_services = CustomerService.objects.filter(
            name__icontains=customer_name
        ).select_related(
            'vpn__vpws_group__mpls_config__equipment'
        ).prefetch_related('vpn__customer_services')
        
        if not customer_services.exists():
            return JsonResponse({'error': f'Cliente "{customer_name}" não encontrado'}, status=400)
        
        # Agrupa por VPN ID para mostrar lado A e B
        vpn_groups = {}
        
        for service in customer_services:
            vpn = service.vpn
            vpn_id = vpn.vpn_id
            
            if vpn_id not in vpn_groups:
                vpn_groups[vpn_id] = {
                    'vpn_id': vpn_id,
                    'description': vpn.description,
                    'encapsulation': vpn.encapsulation,
                    'encapsulation_type': vpn.encapsulation_type,
                    'pw_type': vpn.pw_type,
                    'pw_id': vpn.pw_id,
                    'side_a': None,
                    'side_b': None,
                    'customers': set(),
                    'services': []
                }
            
            # Adiciona cliente e serviço
            vpn_groups[vpn_id]['customers'].add(service.name)
            vpn_groups[vpn_id]['services'].append({
                'name': service.name,
                'type': service.service_type,
                'bandwidth': service.bandwidth
            })
            
            # Busca detalhes da interface de acesso
            access_interface_details = None
            if vpn.access_interface:
                try:
                    interface = Interface.objects.get(
                        mpls_config=vpn.vpws_group.mpls_config,
                        name=vpn.access_interface
                    )
                    access_interface_details = {
                        'name': interface.name,
                        'description': interface.description,
                        'type': interface.interface_type,
                        'speed': interface.speed,
                        'is_customer': interface.is_customer_interface
                    }
                    
                    # Se é LAG, busca membros
                    if interface.interface_type == 'lag':
                        members = list(interface.members.values_list('member_interface_name', flat=True))
                        access_interface_details['lag_members'] = members
                        
                except Interface.DoesNotExist:
                    pass
            
            # Busca equipamento vizinho
            try:
                neighbor_equipment = Equipment.objects.get(ip_address=vpn.neighbor_ip)
                neighbor_hostname = neighbor_equipment.name
            except Equipment.DoesNotExist:
                neighbor_hostname = vpn.neighbor_hostname or 'N/A'
            
            # Cria dados do lado atual
            current_side_data = {
                'equipment': {
                    'hostname': vpn.vpws_group.mpls_config.equipment.name,
                    'loopback_ip': vpn.vpws_group.mpls_config.equipment.ip_address,
                    'location': vpn.vpws_group.mpls_config.equipment.location
                },
                'neighbor': {
                    'hostname': neighbor_hostname,
                    'loopback_ip': vpn.neighbor_ip
                },
                'access_interface': vpn.access_interface,
                'access_interface_details': access_interface_details,
                'vpws_group': vpn.vpws_group.group_name,  # Nome do grupo VPWS do próprio lado
                'encapsulation_details': _extract_encapsulation_details(vpn.encapsulation, vpn.encapsulation_type)
            }
            
            # Busca VPN correspondente no lado oposto
            opposite_vpn = Vpn.objects.filter(
                vpn_id=vpn_id,
                neighbor_ip=vpn.vpws_group.mpls_config.equipment.ip_address
            ).exclude(id=vpn.id).first()
            
            if opposite_vpn:
                # Se encontrou VPN oposta, busca seus detalhes
                try:
                    opposite_equipment = Equipment.objects.get(ip_address=opposite_vpn.vpws_group.mpls_config.equipment.ip_address)
                    opposite_interface_details = None
                    
                    if opposite_vpn.access_interface:
                        try:
                            opposite_interface = Interface.objects.get(
                                mpls_config=opposite_vpn.vpws_group.mpls_config,
                                name=opposite_vpn.access_interface
                            )
                            opposite_interface_details = {
                                'name': opposite_interface.name,
                                'description': opposite_interface.description,
                                'type': opposite_interface.interface_type,
                                'speed': opposite_interface.speed,
                                'is_customer': opposite_interface.is_customer_interface
                            }
                            
                            if opposite_interface.interface_type == 'lag':
                                members = list(opposite_interface.members.values_list('member_interface_name', flat=True))
                                opposite_interface_details['lag_members'] = members
                                
                        except Interface.DoesNotExist:
                            pass
                    
                    opposite_side_data = {
                        'equipment': {
                            'hostname': opposite_equipment.name,
                            'loopback_ip': opposite_equipment.ip_address,
                            'location': opposite_equipment.location
                        },
                        'neighbor': {
                            'hostname': vpn.vpws_group.mpls_config.equipment.name,
                            'loopback_ip': vpn.vpws_group.mpls_config.equipment.ip_address
                        },
                        'access_interface': opposite_vpn.access_interface,
                        'access_interface_details': opposite_interface_details,
                        'vpws_group': opposite_vpn.vpws_group.group_name,  # Nome do grupo VPWS do próprio lado
                        'encapsulation_details': _extract_encapsulation_details(opposite_vpn.encapsulation, opposite_vpn.encapsulation_type)
                    }
                    
                    # Determina qual é lado A e qual é lado B baseado no último octeto do IP loopback
                    current_last_octet = int(vpn.vpws_group.mpls_config.equipment.ip_address.split('.')[-1])
                    opposite_last_octet = int(opposite_equipment.ip_address.split('.')[-1])
                    
                    if current_last_octet < opposite_last_octet:
                        # Lado atual é A, oposto é B
                        vpn_groups[vpn_id]['side_a'] = current_side_data
                        vpn_groups[vpn_id]['side_b'] = opposite_side_data
                    else:
                        # Lado oposto é A, atual é B
                        vpn_groups[vpn_id]['side_a'] = opposite_side_data
                        vpn_groups[vpn_id]['side_b'] = current_side_data
                        
                except Equipment.DoesNotExist:
                    # Se não encontrou equipamento oposto, usa apenas o atual
                    current_last_octet = int(vpn.vpws_group.mpls_config.equipment.ip_address.split('.')[-1])
                    neighbor_last_octet = int(vpn.neighbor_ip.split('.')[-1])
                    
                    if current_last_octet < neighbor_last_octet:
                        vpn_groups[vpn_id]['side_a'] = current_side_data
                    else:
                        vpn_groups[vpn_id]['side_b'] = current_side_data
            else:
                # Se não encontrou VPN oposta, usa apenas o atual
                current_last_octet = int(vpn.vpws_group.mpls_config.equipment.ip_address.split('.')[-1])
                neighbor_last_octet = int(vpn.neighbor_ip.split('.')[-1])
                
                if current_last_octet < neighbor_last_octet:
                    vpn_groups[vpn_id]['side_a'] = current_side_data
                else:
                    vpn_groups[vpn_id]['side_b'] = current_side_data
        
        # Converte sets para listas para serialização JSON
        for vpn_data in vpn_groups.values():
            vpn_data['customers'] = list(vpn_data['customers'])
        
        results = list(vpn_groups.values())
        
        # Registra auditoria da geração do relatório de cliente
        log_audit_action(
            user=request.user,
            action='report_export',
            description=f'Relatório de cliente gerado: {customer_name}',
            request=request,
            search_query=customer_name,
            results_count=len(results),
            additional_data={
                'report_type': 'customer_report',
                'customer_name': customer_name,
                'total_vpns': len(results)
            }
        )
        
        return JsonResponse({
            'customer_name': customer_name,
            'total_vpns': len(results),
            'results': results
        })
        
    except Exception as e:
        import traceback
        print(f"Erro na view customer_report: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'error': f'Erro interno do servidor: {str(e)}'
        }, status=500)


@require_mfa
def customer_report_excel(request):
    """Exporta relatório de cliente para Excel"""
    customer_name = request.GET.get('customer')
    if not customer_name:
        return JsonResponse({'error': 'Nome do cliente é obrigatório'}, status=400)
    
    try:
        # Reutiliza a lógica do customer_report para obter os dados
        customer_services = CustomerService.objects.filter(
            name__icontains=customer_name
        ).select_related(
            'vpn__vpws_group__mpls_config__equipment'
        ).prefetch_related('vpn__customer_services')
        
        if not customer_services.exists():
            return JsonResponse({'error': f'Cliente "{customer_name}" não encontrado'}, status=400)
        
        # Agrupa por VPN ID para mostrar lado A e B
        vpn_groups = {}
        
        for service in customer_services:
            vpn = service.vpn
            vpn_id = vpn.vpn_id
            
            if vpn_id not in vpn_groups:
                vpn_groups[vpn_id] = {
                    'vpn_id': vpn_id,
                    'description': vpn.description,
                    'encapsulation': vpn.encapsulation,
                    'encapsulation_type': vpn.encapsulation_type,
                    'pw_type': vpn.pw_type,
                    'pw_id': vpn.pw_id,
                    'side_a': None,
                    'side_b': None,
                    'customers': set(),
                    'services': []
                }
            
            # Adiciona cliente e serviço
            vpn_groups[vpn_id]['customers'].add(service.name)
            vpn_groups[vpn_id]['services'].append({
                'name': service.name,
                'type': service.service_type,
                'bandwidth': service.bandwidth
            })
            
            # Busca detalhes da interface de acesso
            access_interface_details = None
            if vpn.access_interface:
                try:
                    interface = Interface.objects.get(
                        mpls_config=vpn.vpws_group.mpls_config,
                        name=vpn.access_interface
                    )
                    access_interface_details = {
                        'name': interface.name,
                        'description': interface.description,
                        'type': interface.interface_type,
                        'speed': interface.speed,
                        'is_customer': interface.is_customer_interface
                    }
                    
                    # Se é LAG, busca membros
                    if interface.interface_type == 'lag':
                        members = list(interface.members.values_list('member_interface_name', flat=True))
                        access_interface_details['lag_members'] = members
                        
                except Interface.DoesNotExist:
                    pass
            
            # Busca equipamento vizinho
            try:
                neighbor_equipment = Equipment.objects.get(ip_address=vpn.neighbor_ip)
                neighbor_hostname = neighbor_equipment.name
            except Equipment.DoesNotExist:
                neighbor_hostname = vpn.neighbor_hostname or 'N/A'
            
            # Cria dados do lado atual
            current_side_data = {
                'equipment': {
                    'hostname': vpn.vpws_group.mpls_config.equipment.name,
                    'loopback_ip': vpn.vpws_group.mpls_config.equipment.ip_address,
                    'location': vpn.vpws_group.mpls_config.equipment.location
                },
                'neighbor': {
                    'hostname': neighbor_hostname,
                    'loopback_ip': vpn.neighbor_ip
                },
                'access_interface': vpn.access_interface,
                'access_interface_details': access_interface_details,
                'vpws_group': vpn.vpws_group.group_name,
                'encapsulation_details': _extract_encapsulation_details(vpn.encapsulation, vpn.encapsulation_type)
            }
            
            # Busca VPN correspondente no lado oposto (mesmo código do customer_report)
            opposite_vpn = Vpn.objects.filter(
                vpn_id=vpn_id,
                neighbor_ip=vpn.vpws_group.mpls_config.equipment.ip_address
            ).exclude(id=vpn.id).first()
            
            if opposite_vpn:
                try:
                    opposite_equipment = Equipment.objects.get(ip_address=opposite_vpn.vpws_group.mpls_config.equipment.ip_address)
                    opposite_interface_details = None
                    
                    if opposite_vpn.access_interface:
                        try:
                            opposite_interface = Interface.objects.get(
                                mpls_config=opposite_vpn.vpws_group.mpls_config,
                                name=opposite_vpn.access_interface
                            )
                            opposite_interface_details = {
                                'name': opposite_interface.name,
                                'description': opposite_interface.description,
                                'type': opposite_interface.interface_type,
                                'speed': opposite_interface.speed,
                                'is_customer': opposite_interface.is_customer_interface
                            }
                            
                            if opposite_interface.interface_type == 'lag':
                                members = list(opposite_interface.members.values_list('member_interface_name', flat=True))
                                opposite_interface_details['lag_members'] = members
                                
                        except Interface.DoesNotExist:
                            pass
                    
                    opposite_side_data = {
                        'equipment': {
                            'hostname': opposite_equipment.name,
                            'loopback_ip': opposite_equipment.ip_address,
                            'location': opposite_equipment.location
                        },
                        'neighbor': {
                            'hostname': vpn.vpws_group.mpls_config.equipment.name,
                            'loopback_ip': vpn.vpws_group.mpls_config.equipment.ip_address
                        },
                        'access_interface': opposite_vpn.access_interface,
                        'access_interface_details': opposite_interface_details,
                        'vpws_group': opposite_vpn.vpws_group.group_name,
                        'encapsulation_details': _extract_encapsulation_details(opposite_vpn.encapsulation, opposite_vpn.encapsulation_type)
                    }
                    
                    # Determina qual é lado A e qual é lado B baseado no último octeto do IP loopback
                    current_last_octet = int(vpn.vpws_group.mpls_config.equipment.ip_address.split('.')[-1])
                    opposite_last_octet = int(opposite_equipment.ip_address.split('.')[-1])
                    
                    if current_last_octet < opposite_last_octet:
                        vpn_groups[vpn_id]['side_a'] = current_side_data
                        vpn_groups[vpn_id]['side_b'] = opposite_side_data
                    else:
                        vpn_groups[vpn_id]['side_a'] = opposite_side_data
                        vpn_groups[vpn_id]['side_b'] = current_side_data
                        
                except Equipment.DoesNotExist:
                    current_last_octet = int(vpn.vpws_group.mpls_config.equipment.ip_address.split('.')[-1])
                    neighbor_last_octet = int(vpn.neighbor_ip.split('.')[-1])
                    
                    if current_last_octet < neighbor_last_octet:
                        vpn_groups[vpn_id]['side_a'] = current_side_data
                    else:
                        vpn_groups[vpn_id]['side_b'] = current_side_data
            else:
                current_last_octet = int(vpn.vpws_group.mpls_config.equipment.ip_address.split('.')[-1])
                neighbor_last_octet = int(vpn.neighbor_ip.split('.')[-1])
                
                if current_last_octet < neighbor_last_octet:
                    vpn_groups[vpn_id]['side_a'] = current_side_data
                else:
                    vpn_groups[vpn_id]['side_b'] = current_side_data
        
        # Converte sets para listas
        for vpn_data in vpn_groups.values():
            vpn_data['customers'] = list(vpn_data['customers'])
        
        results = list(vpn_groups.values())
        
        # Cria o Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"Relatório {customer_name}"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        subheader_font = Font(bold=True, color="000000")
        subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Cabeçalho principal
        ws['A1'] = f"Relatório de Cliente: {customer_name}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = center_alignment
        ws.merge_cells('A1:L1')
        
        ws['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A2'].font = Font(italic=True)
        ws['A2'].alignment = center_alignment
        ws.merge_cells('A2:L2')
        
        ws['A3'] = f"Total de VPNs: {len(results)}"
        ws['A3'].font = Font(bold=True)
        ws['A3'].alignment = center_alignment
        ws.merge_cells('A3:L3')
        
        row = 5
        
        # Cabeçalhos das colunas
        headers = [
            'VPN ID', 'Descrição', 'PW Type', 'PW ID', 'Lado A - Equipamento', 
            'Lado A - Loopback', 'Lado A - Interface', 'Lado A - Encapsulation',
            'Lado B - Equipamento', 'Lado B - Loopback', 'Lado B - Interface', 
            'Lado B - Encapsulation'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        row += 1
        
        # Dados das VPNs
        for vpn_data in results:
            side_a = vpn_data.get('side_a', {})
            side_b = vpn_data.get('side_b', {})
            
            # Dados do lado A
            side_a_equipment = side_a.get('equipment', {}).get('hostname', 'N/A') if side_a else 'N/A'
            side_a_loopback = side_a.get('equipment', {}).get('loopback_ip', 'N/A') if side_a else 'N/A'
            side_a_interface = side_a.get('access_interface', 'N/A') if side_a else 'N/A'
            side_a_encap = _format_encapsulation_for_excel(side_a.get('encapsulation_details') if side_a else None)
            
            # Dados do lado B
            side_b_equipment = side_b.get('equipment', {}).get('hostname', 'N/A') if side_b else 'N/A'
            side_b_loopback = side_b.get('equipment', {}).get('loopback_ip', 'N/A') if side_b else 'N/A'
            side_b_interface = side_b.get('access_interface', 'N/A') if side_b else 'N/A'
            side_b_encap = _format_encapsulation_for_excel(side_b.get('encapsulation_details') if side_b else None)
            
            data_row = [
                vpn_data['vpn_id'],
                vpn_data.get('description', 'N/A'),
                vpn_data.get('pw_type', 'N/A'),
                vpn_data.get('pw_id', 'N/A'),
                side_a_equipment,
                side_a_loopback,
                side_a_interface,
                side_a_encap,
                side_b_equipment,
                side_b_loopback,
                side_b_interface,
                side_b_encap
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = center_alignment
            
            row += 1
        
        # Ajusta largura das colunas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        # Registra auditoria da exportação
        log_audit_action(
            user=request.user,
            action='report_export',
            description=f'Exportação Excel do relatório de cliente: {customer_name}',
            request=request,
            search_query=customer_name,
            export_format='xlsx',
            results_count=len(results),
            additional_data={
                'report_type': 'customer_report_excel',
                'customer_name': customer_name,
                'total_vpns': len(results)
            }
        )
        
        # Cria resposta HTTP com Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="relatorio_cliente_{customer_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        import traceback
        print(f"Erro na exportação Excel: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'error': f'Erro ao gerar Excel: {str(e)}'
        }, status=500)


def _format_encapsulation_for_excel(encapsulation_details):
    """Formata detalhes de encapsulamento para exibição no Excel"""
    if not encapsulation_details:
        return 'N/A'
    
    encap_type = encapsulation_details.get('type', '')
    vlans = encapsulation_details.get('vlans', [])
    
    if not vlans:
        # Se não há VLANs processadas, usa o raw
        raw = encapsulation_details.get('raw', '')
        if raw:
            return f"Encapsulation Dot1q: {raw}"
        return 'N/A'
    
    # Formata baseado no tipo
    if encap_type == 'qinq':
        # Para QINQ, procura outer e inner VLANs
        outer_vlan = None
        inner_vlan = None
        
        for vlan in vlans:
            if vlan.get('type') == 'outer':
                outer_vlan = vlan.get('vlan')
            elif vlan.get('type') == 'inner':
                inner_vlan = vlan.get('vlan')
        
        if outer_vlan and inner_vlan:
            return f"Encapsulation Dot1q: {outer_vlan}.{inner_vlan}"
        elif outer_vlan:
            return f"Encapsulation Dot1q: {outer_vlan}"
        
    elif encap_type == 'vlan_tagged':
        # Para VLAN tagged, usa a primeira VLAN
        if vlans and vlans[0].get('vlan'):
            return f"Encapsulation Dot1q: {vlans[0]['vlan']}"
    
    # Fallback para formato genérico
    vlan_numbers = [str(vlan.get('vlan', '')) for vlan in vlans if vlan.get('vlan')]
    if vlan_numbers:
        if len(vlan_numbers) == 1:
            return f"Encapsulation Dot1q: {vlan_numbers[0]}"
        else:
            return f"Encapsulation Dot1q: {'.'.join(vlan_numbers)}"
    
    # Se tudo falhar, usa raw
    raw = encapsulation_details.get('raw', '')
    if raw:
        return f"Encapsulation Dot1q: {raw}"
    
    return 'N/A'


def _extract_encapsulation_details(encapsulation, encapsulation_type):
    """Extrai detalhes de encapsulamento para exibição no relatório"""
    if not encapsulation:
        return None
    
    details = {
        'type': encapsulation_type,
        'raw': encapsulation,
        'vlans': []
    }
    
    if encapsulation_type == 'qinq':
        # Para QINQ, tenta extrair as VLANs outer e inner
        # Exemplos de formatos: "100:200", "100-200", "100/200", "100.200"
        import re
        
        # Padrões comuns para QINQ
        patterns = [
            r'(\d+)[:.\-/](\d+)',  # 100:200, 100.200, 100-200, 100/200
            r'(\d+)\s+(\d+)',      # 100 200
            r'outer\s*(\d+).*inner\s*(\d+)',  # outer 100 inner 200
            r'(\d+)\s*to\s*(\d+)', # 100 to 200
        ]
        
        for pattern in patterns:
            match = re.search(pattern, encapsulation, re.IGNORECASE)
            if match:
                outer_vlan = match.group(1)
                inner_vlan = match.group(2)
                details['vlans'] = [
                    {'type': 'outer', 'vlan': outer_vlan},
                    {'type': 'inner', 'vlan': inner_vlan}
                ]
                break
        
        # Se não encontrou padrão específico, tenta extrair números
        if not details['vlans']:
            vlan_numbers = re.findall(r'\d+', encapsulation)
            if len(vlan_numbers) >= 2:
                details['vlans'] = [
                    {'type': 'outer', 'vlan': vlan_numbers[0]},
                    {'type': 'inner', 'vlan': vlan_numbers[1]}
                ]
    
    elif encapsulation_type == 'vlan_tagged':
        # Para VLAN tagged, tenta extrair o número da VLAN
        import re
        vlan_match = re.search(r'(\d+)', encapsulation)
        if vlan_match:
            details['vlans'] = [
                {'type': 'vlan', 'vlan': vlan_match.group(1)}
            ]
    
    return details


@require_mfa
def search_view(request):
    query = request.GET.get('q', '').strip()
    equipment_filter = request.GET.get('equipment', '')
    location_filter = request.GET.get('location', '')
    raw_service_type_filter = request.GET.get('service_type', '')
    service_type_filter = raw_service_type_filter.strip().lower()

    results = []

    if query or equipment_filter or location_filter or service_type_filter:
        # Base da busca: começar pelas VPNs, trazendo equipamento e serviços
        vpn_qs = (
            Vpn.objects.select_related('vpws_group__mpls_config__equipment')
            .prefetch_related('customer_services')
        )

        client_name_filter = None
        if query:
            if query.isdigit():
                vpn_qs = vpn_qs.filter(vpn_id=int(query))
            else:
                # Se parecer nome de cliente, filtra pelos serviços; senão, campos gerais
                if re.match(r'^[A-Za-z][A-Za-z0-9&\- _\.]{2,}$', query):
                    vpn_qs = vpn_qs.filter(customer_services__name__icontains=query)
                    client_name_filter = query
                else:
                    vpn_qs = vpn_qs.filter(
                        Q(vpws_group__mpls_config__equipment__name__icontains=query) |
                        Q(vpws_group__mpls_config__equipment__location__icontains=query) |
                        Q(description__icontains=query) |
                        Q(access_interface__icontains=query) |
                        Q(encapsulation__icontains=query)
                    )

        if equipment_filter:
            vpn_qs = vpn_qs.filter(
                vpws_group__mpls_config__equipment__name__icontains=equipment_filter
            )

        if location_filter:
            vpn_qs = vpn_qs.filter(
                vpws_group__mpls_config__equipment__location__icontains=location_filter
            )

        if service_type_filter:
            allowed_service_types = {choice[0] for choice in CustomerService.SERVICE_TYPE_CHOICES}
            if service_type_filter in allowed_service_types:
                vpn_qs = vpn_qs.filter(customer_services__service_type=service_type_filter)
            else:
                service_type_filter = ''

        vpn_qs = vpn_qs.distinct()

        for vpn in vpn_qs:
            equipment = vpn.vpws_group.mpls_config.equipment
            vpws_group = vpn.vpws_group

            # Equipamento remoto
            try:
                neighbor_equipment = Equipment.objects.get(ip_address=vpn.neighbor_ip)
            except Equipment.DoesNotExist:
                neighbor_equipment = None

            # Detalhes da interface
            access_interface_details = None
            if vpn.access_interface:
                try:
                    interface = Interface.objects.get(
                        mpls_config=vpws_group.mpls_config,
                        name=vpn.access_interface
                    )
                    access_interface_details = {
                        'name': interface.name,
                        'description': interface.description,
                        'type': interface.interface_type,
                        'speed': interface.speed
                    }
                    if interface.interface_type == 'lag':
                        members = list(interface.members.values_list('member_interface_name', flat=True))
                        access_interface_details['lag_members'] = members
                except Interface.DoesNotExist:
                    pass

            # Busca interface do lado oposto (se existir)
            opposite_interface_details = None
            try:
                # Busca VPN com mesmo ID no equipamento vizinho
                opposite_vpn = Vpn.objects.filter(
                    vpn_id=vpn.vpn_id,
                    neighbor_ip=equipment.ip_address
                ).first()
                
                if opposite_vpn and opposite_vpn.access_interface:
                    opposite_interface = Interface.objects.filter(
                        mpls_config__equipment=neighbor_equipment,
                        name=opposite_vpn.access_interface
                    ).first()
                    
                    if opposite_interface:
                        opposite_interface_details = {
                            'name': opposite_interface.name,
                            'description': opposite_interface.description,
                            'type': opposite_interface.interface_type,
                            'speed': opposite_interface.speed
                        }
                        if opposite_interface.interface_type == 'lag':
                            members = list(opposite_interface.members.values_list('member_interface_name', flat=True))
                            opposite_interface_details['lag_members'] = members
            except:
                pass

            # Seleciona serviços de cliente associados a esta VPN
            cs_list = list(vpn.customer_services.all())
            if client_name_filter:
                cs_list = [cs for cs in cs_list if client_name_filter.lower() in cs.name.lower()]
            if service_type_filter:
                cs_list = [cs for cs in cs_list if cs.service_type == service_type_filter]

            # Se não houver cliente filtrado, ainda mostramos a VPN (cliente vazio)
            if not cs_list:
                results.append({
                    'customer_name': '',
                    'equipment_name': equipment.name,
                    'equipment_ip': equipment.ip_address,
                    'location': equipment.location,
                    'service_type': '',
                    'vpn_id': vpn.vpn_id,
                    'vpn_description': vpn.description,
                    'neighbor_ip': vpn.neighbor_ip,
                    'neighbor_equipment': neighbor_equipment.name if neighbor_equipment else None,
                    'vpws_group_name': vpws_group.group_name,
                    'encapsulation': vpn.encapsulation,
                    'encapsulation_type': vpn.encapsulation_type,
                    'access_interface': vpn.access_interface,
                    'access_interface_details': access_interface_details,
                    'opposite_interface_details': opposite_interface_details,
                    'pw_type': vpn.pw_type,
                    'pw_id': vpn.pw_id,
                    'last_backup': equipment.last_backup,
                })
            else:
                for service in cs_list:
                    results.append({
                        'customer_name': service.name,
                        'equipment_name': equipment.name,
                        'equipment_ip': equipment.ip_address,
                        'location': equipment.location,
                        'service_type': service.get_service_type_display(),
                        'vpn_id': vpn.vpn_id,
                        'vpn_description': vpn.description,
                        'neighbor_ip': vpn.neighbor_ip,
                        'neighbor_equipment': neighbor_equipment.name if neighbor_equipment else None,
                        'vpws_group_name': vpws_group.group_name,
                        'encapsulation': vpn.encapsulation,
                        'encapsulation_type': vpn.encapsulation_type,
                        'access_interface': vpn.access_interface,
                        'access_interface_details': access_interface_details,
                        'opposite_interface_details': opposite_interface_details,
                        'pw_type': vpn.pw_type,
                        'pw_id': vpn.pw_id,
                        'last_backup': equipment.last_backup,
                    })

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Registra auditoria da busca se há consulta
    if query or equipment_filter or location_filter or service_type_filter:
        search_terms = []
        if query:
            search_terms.append(f"query: '{query}'")
        if equipment_filter:
            search_terms.append(f"equipment: '{equipment_filter}'")
        if location_filter:
            search_terms.append(f"location: '{location_filter}'")
        if service_type_filter:
            search_terms.append(f"service_type: '{service_type_filter}'")
        
        search_description = f"Busca realizada com filtros: {', '.join(search_terms)}"
        AuditLogger.log_search(
            user=request.user,
            request=request,
            search_query=search_description,
            results_count=len(results)
        )

    equipment_options = Equipment.objects.values_list('name', flat=True).distinct()
    location_options = Equipment.objects.exclude(location='').values_list('location', flat=True).distinct()
    service_type_options = CustomerService.SERVICE_TYPE_CHOICES

    context = {
        'query': query,
        'equipment_filter': equipment_filter,
        'location_filter': location_filter,
        'service_type_filter': service_type_filter,
        'page_obj': page_obj,
        'results_count': len(results),
        'equipment_options': equipment_options,
        'location_options': location_options,
        'service_type_options': [choice[0] for choice in service_type_options],
    }

    return render(request, 'mpls_analyzer/search.html', context)


@require_mfa
def equipment_detail(request, equipment_id):
    try:
        equipment = Equipment.objects.get(id=equipment_id)
        
        # Registra auditoria da visualização
        AuditLogger.log_equipment_view(
            user=request.user,
            request=request,
            equipment_id=equipment_id,
            equipment_name=equipment.name
        )
        
        # Configurações MPLS mais recentes
        latest_config = equipment.mpls_configs.first()
        
        # Serviços de clientes
        customer_services = CustomerService.objects.filter(
            vpn__vpws_group__mpls_config__equipment=equipment
        ).distinct()
        
        # VPNs
        vpns = Vpn.objects.filter(
            vpws_group__mpls_config__equipment=equipment
        ).select_related('vpws_group')
        
        context = {
            'equipment': equipment,
            'latest_config': latest_config,
            'customer_services': customer_services,
            'vpns': vpns,
        }
        
        return render(request, 'mpls_analyzer/equipment_detail.html', context)
        
    except Equipment.DoesNotExist:
        messages.error(request, 'Equipamento não encontrado.')
        return redirect('search')


@require_mfa
@ratelimit(key='user', rate='3/m', method='POST', block=True)
def update_database_view(request):
    if request.method == 'POST':
        try:
            # Captura credenciais do formulário
            username = request.POST.get('device_username', '').strip()
            password = request.POST.get('device_password', '').strip()
            
            if not username or not password:
                messages.error(request, 'Login e senha dos equipamentos são obrigatórios.')
                return redirect('update_database')
            
            # Executa scan da rede e backup dos equipamentos
            import subprocess
            import os
            from django.conf import settings
            
            # Define diretório base do projeto
            base_dir = settings.BASE_DIR
            
            # Caminhos dos scripts dentro do app
            scripts_dir = os.path.join(base_dir, 'mpls_analyzer', 'scripts')
            scan_script_path = os.path.join(scripts_dir, 'scan-network.py')
            backup_script_path = os.path.join(scripts_dir, 'easy-bkp-optimized.py')
            
            # Executa scan-network.py com credenciais
            scan_env = os.environ.copy()
            scan_env['DEVICE_USERNAME'] = username
            scan_env['DEVICE_PASSWORD'] = password
            
            messages.info(request, 'Iniciando escaneamento da rede...')
            scan_result = subprocess.run(
                ['python', scan_script_path],
                env=scan_env,
                capture_output=True,
                text=True,
                timeout=300  # 5 minutos timeout
            )
            
            if scan_result.returncode != 0:
                raise Exception(f'Erro no escaneamento: {scan_result.stderr}')
            
            messages.success(request, 'Escaneamento da rede concluído!')
            
            # Executa easy-bkp-simplified.py com credenciais
            backup_env = os.environ.copy()
            backup_env['DEVICE_USERNAME'] = username
            backup_env['DEVICE_PASSWORD'] = password
            
            messages.info(request, 'Iniciando backup dos equipamentos...')
            backup_result = subprocess.run(
                ['python', backup_script_path],
                env=backup_env,
                capture_output=True,
                text=True,
                timeout=1800  # 30 minutos timeout
            )
            
            if backup_result.returncode != 0:
                raise Exception(f'Erro no backup: {backup_result.stderr}')
            
            messages.success(request, 'Backup dos equipamentos concluído!')
            
            # Executa processamento da base de dados
            from django.core.management import call_command
            from datetime import datetime
            
            # Define o diretório de backup criado pelo script
            backup_date = datetime.now().strftime("%Y-%m-%d")
            backup_dir = os.path.join(scripts_dir, f"backup_{backup_date}")
            
            messages.info(request, 'Processando dados para a base de dados...')
            
            # Primeiro, tenta processar normalmente
            initial_device_count = len([f for f in os.listdir(backup_dir) if f.endswith('.json')])
            
            call_command('update_database', 
                        backup_dir=backup_dir,
                        user=request.user.username)
            
            # Verifica se há falhas de processamento que precisam de correção JSON
            from mpls_analyzer.models import Equipment
            processed_count = Equipment.objects.count()
            
            # Se há diferença significativa, executa correção automática
            if initial_device_count < 103:  # Total esperado de equipamentos
                messages.info(request, f'Detectados {103 - initial_device_count} equipamentos com falha de JSON. Iniciando correção automática...')
                
                # Executa correção inteligente de JSON
                smart_fix_script = os.path.join(scripts_dir, 'smart-json-fix.py')
                fix_result = subprocess.run(
                    ['python', smart_fix_script],
                    capture_output=True,
                    text=True,
                    timeout=300
                )
                
                if fix_result.returncode == 0:
                    messages.success(request, 'Correção de JSON concluída! Reprocessando dados...')
                    
                    # Reprocessa dados após correção
                    call_command('update_database', 
                                backup_dir=backup_dir,
                                user=request.user.username)
                    
                    final_count = Equipment.objects.count()
                    messages.success(request, f'Base de dados atualizada! Total de equipamentos processados: {final_count}')
                else:
                    messages.warning(request, 'Correção automática falhou. Use a interface de correção manual se necessário.')
            else:
                messages.success(request, 'Base de dados atualizada com sucesso!')
            
        except subprocess.TimeoutExpired:
            messages.error(request, 'Timeout: O processo demorou mais que o esperado.')
        except Exception as e:
            messages.error(request, f'Erro ao atualizar base de dados: {str(e)}')
    
    # Lista logs recentes
    logs = BackupProcessLog.objects.all()[:10]
    
    context = {
        'logs': logs,
    }
    
    return render(request, 'mpls_analyzer/update_database.html', context)


@require_mfa
def fix_malformed_json(request):
    """View para corrigir JSONs malformados - trabalha com dados já coletados"""
    if request.method == 'POST':
        try:
            import subprocess
            import os
            from django.conf import settings
            
            # Define diretório base do projeto
            base_dir = settings.BASE_DIR
            scripts_dir = os.path.join(base_dir, 'mpls_analyzer', 'scripts')
            smart_fix_script = os.path.join(scripts_dir, 'smart-json-fix.py')
            
            if not os.path.exists(smart_fix_script):
                messages.error(request, 'Script de correção inteligente não encontrado.')
                return redirect('fix_malformed_json')
            
            messages.info(request, 'Executando correção inteligente de JSONs malformados...')
            fix_result = subprocess.run(
                ['python', smart_fix_script],
                capture_output=True,
                text=True,
                timeout=600  # 10 minutos timeout
            )
            
            if fix_result.returncode != 0:
                messages.error(request, f'Erro na correção: {fix_result.stderr}')
            else:
                # Processa a saída para mostrar resultados
                output_lines = fix_result.stdout.split('\n')
                success_count = 0
                failed_count = 0
                
                for line in output_lines:
                    if '✅ Corrigidos com sucesso:' in line:
                        success_count = int(line.split(':')[1].strip())
                    elif '❌ Ainda com problemas:' in line or '❌ Falha:' in line:
                        failed_count = int(line.split(':')[1].strip())
                
                if success_count > 0:
                    messages.success(request, f'Correção concluída! {success_count} equipamentos corrigidos com sucesso.')
                    if failed_count > 0:
                        messages.warning(request, f'{failed_count} equipamentos ainda apresentam problemas.')
                else:
                    messages.warning(request, 'Nenhum equipamento foi corrigido. Verifique os logs para mais detalhes.')
                
                # Mostra saída completa em debug
                messages.info(request, f'Saída do script: {fix_result.stdout[-500:]}...')  # Últimos 500 chars
            
        except subprocess.TimeoutExpired:
            messages.error(request, 'Timeout: O processo de correção demorou mais que o esperado.')
        except Exception as e:
            messages.error(request, f'Erro ao executar correção: {str(e)}')
    
    # Lista scripts de correção disponíveis
    import os
    from django.conf import settings
    
    scripts_dir = os.path.join(settings.BASE_DIR, 'mpls_analyzer', 'scripts')
    available_scripts = []
    
    fix_scripts = [
        'detailed-debug.py',
        'precise-fix.py',
        'manual-fix-devices.py',
        'ultimate-fix.py', 
        'simple-fix.py',
        'fix-failed-devices.py',
        'debug-json-error.py'
    ]
    
    for script in fix_scripts:
        script_path = os.path.join(scripts_dir, script)
        if os.path.exists(script_path):
            available_scripts.append({
                'name': script,
                'display_name': script.replace('-', ' ').replace('.py', '').title(),
                'path': script_path
            })
    
    context = {
        'available_scripts': available_scripts,
    }
    
    return render(request, 'mpls_analyzer/fix_malformed_json.html', context)


@require_mfa
def api_search(request):
    """API endpoint para busca em tempo real (AJAX)"""
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    # Se for numérico, prioriza busca por VPN ID
    if query.isdigit():
        vpn_qs = (
            Vpn.objects.filter(vpn_id=int(query))
            .select_related('vpws_group__mpls_config__equipment')
            .prefetch_related('customer_services')
        )
        vpn_results = []
        for v in vpn_qs[:10]:
            equip = v.vpws_group.mpls_config.equipment
            customers = [cs.name for cs in v.customer_services.all()]
            vpn_results.append({
                'type': 'vpn',
                'vpn_id': v.vpn_id,
                'equipment_name': equip.name,
                'equipment_id': equip.id,
                'loopback_ip': equip.ip_address,
                'neighbor_ip': v.neighbor_ip,
                'neighbor_hostname': v.neighbor_hostname,
                'access_interface': v.access_interface,
                'encapsulation': v.encapsulation,
                'description': v.description,
                'group_name': v.vpws_group.group_name,
                'customers': customers,
            })
        if vpn_results:
            return JsonResponse({'results': vpn_results})

    # Busca por nome de cliente: retornar diretamente as VPNs vinculadas ao cliente
    customer_services = (
        CustomerService.objects.filter(
            Q(name__iexact=query) | Q(name__icontains=query)
        )
        .select_related('vpn__vpws_group__mpls_config__equipment')
        .prefetch_related('vpn__customer_services')
    )

    vpn_results = []
    seen_vpn_ids = set()
    for service in customer_services:
        v = service.vpn
        if v.id in seen_vpn_ids:
            continue
        seen_vpn_ids.add(v.id)
        equip = v.vpws_group.mpls_config.equipment
        customers = [cs.name for cs in v.customer_services.all()]
        vpn_results.append({
            'type': 'vpn',
            'vpn_id': v.vpn_id,
            'equipment_name': equip.name,
            'equipment_id': equip.id,
            'loopback_ip': equip.ip_address,
            'neighbor_ip': v.neighbor_ip,
            'neighbor_hostname': v.neighbor_hostname,
            'access_interface': v.access_interface,
            'encapsulation': v.encapsulation,
            'description': v.description,
            'group_name': v.vpws_group.group_name,
            'customers': customers,
        })
        if len(vpn_results) >= 10:
            break

    return JsonResponse({'results': vpn_results})


# Funções auxiliares para verificação de permissões
def is_admin(user):
    """Verifica se o usuário é administrador"""
    return user.is_authenticated and (user.is_superuser or hasattr(user, 'profile') and user.profile.is_admin)


def check_mfa_requirement(user):
    """Verifica se o usuário precisa de MFA baseado no seu profile"""
    if not hasattr(user, 'profile'):
        return True  # Por padrão, requer MFA
    return user.profile.require_mfa


# Views administrativas
@user_passes_test(is_admin)
def admin_panel(request):
    """Painel administrativo principal"""
    users_count = User.objects.count()
    active_users = User.objects.filter(is_active=True).count()
    users_with_mfa = User.objects.filter(profile__require_mfa=True).count()
    
    recent_users = User.objects.select_related('profile').order_by('-date_joined')[:10]
    
    context = {
        'users_count': users_count,
        'active_users': active_users,
        'users_with_mfa': users_with_mfa,
        'recent_users': recent_users,
    }
    
    return render(request, 'mpls_analyzer/admin/panel.html', context)


@user_passes_test(is_admin)
def admin_users_list(request):
    """Lista de usuários para administração"""
    users = User.objects.select_related('profile').all()
    
    # Filtros
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    mfa_filter = request.GET.get('mfa', '')
    
    if search:
        users = users.filter(
            Q(username__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(email__icontains=search)
        )
    
    if status_filter == 'active':
        users = users.filter(is_active=True)
    elif status_filter == 'inactive':
        users = users.filter(is_active=False)
    
    if mfa_filter == 'required':
        users = users.filter(profile__require_mfa=True)
    elif mfa_filter == 'optional':
        users = users.filter(profile__require_mfa=False)
    
    # Paginação
    paginator = Paginator(users, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'status_filter': status_filter,
        'mfa_filter': mfa_filter,
    }
    
    return render(request, 'mpls_analyzer/admin/users_list.html', context)


@user_passes_test(is_admin)
def admin_user_detail(request, user_id):
    """Detalhes e edição de usuário"""
    user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'toggle_active':
            user.is_active = not user.is_active
            user.save()
            status = "ativado" if user.is_active else "desativado"
            messages.success(request, f'Usuário {status} com sucesso!')
            log_security_event(f'USER_{status.upper()}', user=request.user, 
                             details=f'Target user: {user.username}')
        
        elif action == 'toggle_mfa':
            user.profile.require_mfa = not user.profile.require_mfa
            user.profile.save()
            status = "obrigatório" if user.profile.require_mfa else "opcional"
            messages.success(request, f'MFA definido como {status}!')
            log_security_event(f'MFA_REQUIREMENT_CHANGED', user=request.user,
                             details=f'Target user: {user.username}, MFA required: {user.profile.require_mfa}')
        
        elif action == 'toggle_admin':
            user.profile.is_admin = not user.profile.is_admin
            user.profile.save()
            status = "administrador" if user.profile.is_admin else "usuário comum"
            messages.success(request, f'Usuário definido como {status}!')
            log_security_event(f'ADMIN_PRIVILEGES_CHANGED', user=request.user,
                             details=f'Target user: {user.username}, Is admin: {user.profile.is_admin}')
        
        elif action == 'unlock_account':
            user.profile.unlock_account()
            messages.success(request, 'Conta desbloqueada com sucesso!')
            log_security_event('ACCOUNT_UNLOCKED', user=request.user,
                             details=f'Target user: {user.username}')
        
        elif action == 'reset_mfa':
            # MFA disabled for development
            messages.success(request, 'MFA resetado! (simulado para desenvolvimento)')
            log_security_event('MFA_RESET', user=request.user,
                             details=f'Target user: {user.username}')
        
        elif action == 'reset_password':
            new_password = request.POST.get('new_password')
            if new_password:
                # Valida senha
                errors = validate_password_strength(new_password)
                if errors:
                    for error in errors:
                        messages.error(request, error)
                else:
                    user.set_password(new_password)
                    user.profile.last_password_change = timezone.now()
                    user.save()
                    user.profile.save()
                    messages.success(request, 'Senha alterada com sucesso!')
                    log_security_event('PASSWORD_RESET_BY_ADMIN', user=request.user,
                                     details=f'Target user: {user.username}')
        
        return redirect('admin_user_detail', user_id=user.id)
    
    # MFA disabled for development
    mfa_devices = []
    
    context = {
        'target_user': user,
        'mfa_devices': mfa_devices,
    }
    
    return render(request, 'mpls_analyzer/admin/user_detail.html', context)


@user_passes_test(is_admin)
def admin_create_user(request):
    """Criar novo usuário"""
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        password = request.POST.get('password')
        require_mfa = request.POST.get('require_mfa') == 'on'
        is_admin = request.POST.get('is_admin') == 'on'
        
        # Validações
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Usuário já existe!')
        elif not password:
            messages.error(request, 'Senha é obrigatória!')
        else:
            # Valida senha
            errors = validate_password_strength(password)
            if errors:
                for error in errors:
                    messages.error(request, error)
            else:
                # Cria usuário
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    password=password
                )
                
                # Configura profile
                user.profile.require_mfa = require_mfa
                user.profile.is_admin = is_admin
                user.profile.last_password_change = timezone.now()
                user.profile.save()
                
                messages.success(request, f'Usuário {username} criado com sucesso!')
                log_security_event('USER_CREATED', user=request.user,
                                 details=f'New user: {username}, MFA required: {require_mfa}, Is admin: {is_admin}')
                
                return redirect('admin_users_list')
    
    return render(request, 'mpls_analyzer/admin/create_user.html')


# Modificar a view de busca para usar o novo sistema
@require_mfa
def advanced_search_view(request):
    """Busca avançada com full-text search"""
    query = request.GET.get('q', '').strip()
    search_type = request.GET.get('type', 'auto')
    
    results = []
    search_engine = None
    
    if query:
        search_engine = AdvancedSearchEngine()
        
        if search_type == 'full_text':
            configs = search_engine.search_full_text(query)
        else:
            configs = smart_search(query, search_type)
        
        # Prepara resultados com highlights
        for config in configs[:50]:  # Limita a 50 resultados
            highlights = search_engine.extract_search_highlights(config.raw_config, query)
            results.append({
                'config': config,
                'highlights': highlights,
                'equipment': config.equipment,
            })
    
    context = {
        'query': query,
        'search_type': search_type,
        'results': results,
        'results_count': len(results),
    }
    
    return render(request, 'mpls_analyzer/advanced_search.html', context)


@require_mfa
def update_status_api(request):
    """API para verificar status do último processo de atualização"""
    try:
        latest_log = BackupProcessLog.objects.latest('started_at')
        
        return JsonResponse({
            'status': latest_log.status,
            'started_at': latest_log.started_at.isoformat() if latest_log.started_at else None,
            'finished_at': latest_log.finished_at.isoformat() if latest_log.finished_at else None,
            'processed_files': latest_log.processed_files or 0,
            'total_files': latest_log.total_files or 0,
            'errors': latest_log.errors,
            'user': latest_log.user.username if latest_log.user else None
        })
    except BackupProcessLog.DoesNotExist:
        return JsonResponse({
            'status': 'no_logs',
            'message': 'Nenhum processo encontrado'
        })


# =============================================================================
# VIEWS DE GERENCIAMENTO DE USUÁRIOS
# =============================================================================

def admin_required(view_func):
    """Decorator para verificar se o usuário é admin"""
    def wrapper(request, *args, **kwargs):
        if not hasattr(request.user, 'profile') or not request.user.profile.is_admin:
            messages.error(request, 'Acesso negado. Apenas administradores podem acessar esta área.')
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


@require_mfa
@admin_required
def user_management(request):
    """View principal para gerenciamento de usuários"""
    users = User.objects.select_related('profile').order_by('username')
    
    context = {
        'users': users,
        'total_users': users.count(),
        'active_users': users.filter(is_active=True).count(),
        'admin_users': users.filter(profile__is_admin=True).count(),
    }
    
    return render(request, 'mpls_analyzer/user_management.html', context)


@require_mfa
@admin_required
def create_user(request):
    """View para criar novo usuário"""
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            try:
                user = form.save()
                log_security_event(
                    'USER_CREATED',
                    f'Usuário {user.username} criado por {request.user.username}',
                    request.user,
                    get_client_ip(request)
                )
                messages.success(request, f'Usuário {user.username} criado com sucesso!')
                return redirect('user_management')
            except Exception as e:
                messages.error(request, f'Erro ao criar usuário: {str(e)}')
    else:
        form = UserRegistrationForm()
    
    context = {
        'form': form,
        'title': 'Criar Novo Usuário'
    }
    
    return render(request, 'mpls_analyzer/create_user.html', context)


@require_mfa
@admin_required
def edit_user(request, user_id):
    """View para editar usuário (apenas admins)"""
    user = get_object_or_404(User, id=user_id)
    profile, created = UserProfile.objects.get_or_create(user=user)
    
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=profile, user=user)
        if form.is_valid():
            try:
                form.save()
                log_security_event(
                    'USER_UPDATED',
                    f'Usuário {user.username} atualizado por {request.user.username}',
                    request.user,
                    get_client_ip(request)
                )
                messages.success(request, f'Usuário {user.username} atualizado com sucesso!')
                return redirect('user_management')
            except Exception as e:
                messages.error(request, f'Erro ao atualizar usuário: {str(e)}')
    else:
        form = UserProfileForm(instance=profile, user=user)
    
    context = {
        'form': form,
        'user_obj': user,
        'title': f'Editar Usuário: {user.username}'
    }
    
    return render(request, 'mpls_analyzer/edit_user.html', context)


@require_mfa
def user_profile(request):
    """View para o usuário editar seu próprio perfil"""
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=profile, user=request.user)
        if form.is_valid():
            try:
                form.save()
                log_security_event(
                    'PROFILE_UPDATED',
                    f'Perfil atualizado por {request.user.username}',
                    request.user,
                    get_client_ip(request)
                )
                messages.success(request, 'Perfil atualizado com sucesso!')
                return redirect('user_profile')
            except Exception as e:
                messages.error(request, f'Erro ao atualizar perfil: {str(e)}')
    else:
        form = UserProfileForm(instance=profile, user=request.user)
    
    context = {
        'form': form,
        'user_obj': request.user,
        'title': 'Meu Perfil'
    }
    
    return render(request, 'mpls_analyzer/user_profile.html', context)


@require_mfa
def change_password(request):
    """View para alterar senha"""
    if request.method == 'POST':
        form = StrongPasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            try:
                user = form.save()
                # Atualiza data da última alteração de senha
                profile, created = UserProfile.objects.get_or_create(user=user)
                profile.last_password_change = timezone.now()
                profile.save()
                
                log_security_event(
                    'PASSWORD_CHANGED',
                    f'Senha alterada por {request.user.username}',
                    request.user,
                    get_client_ip(request)
                )
                messages.success(request, 'Senha alterada com sucesso!')
                return redirect('user_profile')
            except Exception as e:
                messages.error(request, f'Erro ao alterar senha: {str(e)}')
    else:
        form = StrongPasswordChangeForm(request.user)
    
    context = {
        'form': form,
        'title': 'Alterar Senha'
    }
    
    return render(request, 'mpls_analyzer/change_password.html', context)


@require_mfa
@admin_required
def toggle_user_status(request, user_id):
    """API para ativar/desativar usuário"""
    if request.method == 'POST':
        user = get_object_or_404(User, id=user_id)
        
        # Não permite desativar o próprio usuário
        if user == request.user:
            return JsonResponse({
                'success': False,
                'message': 'Você não pode desativar sua própria conta.'
            })
        
        user.is_active = not user.is_active
        user.save()
        
        action = 'ativado' if user.is_active else 'desativado'
        log_security_event(
            'USER_STATUS_CHANGED',
            f'Usuário {user.username} {action} por {request.user.username}',
            request.user,
            get_client_ip(request)
        )
        
        return JsonResponse({
            'success': True,
            'is_active': user.is_active,
            'message': f'Usuário {user.username} {action} com sucesso!'
        })
    
    return JsonResponse({'success': False, 'message': 'Método não permitido'})


@require_mfa
@admin_required
def delete_user(request, user_id):
    """API para excluir usuário"""
    if request.method == 'POST':
        user = get_object_or_404(User, id=user_id)
        
        # Não permite excluir o próprio usuário
        if user == request.user:
            return JsonResponse({
                'success': False,
                'message': 'Você não pode excluir sua própria conta.'
            })
        
        username = user.username
        user.delete()
        
        log_security_event(
            'USER_DELETED',
            f'Usuário {username} excluído por {request.user.username}',
            request.user,
            get_client_ip(request)
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Usuário {username} excluído com sucesso!'
        })
    
    return JsonResponse({'success': False, 'message': 'Método não permitido'})


@require_mfa
def setup_mfa(request):
    """View para configurar MFA pela primeira vez"""
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    if request.method == 'POST':
        token = request.POST.get('token')
        secret = request.session.get('mfa_temp_secret')
        
        if not secret:
            messages.error(request, 'Sessão expirada. Tente novamente.')
            return redirect('setup_mfa')
        
        totp = pyotp.TOTP(secret)
        if totp.verify(token):
            profile.mfa_secret = secret
            profile.mfa_enabled = True
            profile.require_mfa = True
            profile.save()
            
            # Remove secret temporário da sessão
            del request.session['mfa_temp_secret']
            
            log_security_event(
                'MFA_ENABLED',
                f'MFA habilitado por {request.user.username}',
                request.user,
                get_client_ip(request)
            )
            
            messages.success(request, 'MFA configurado com sucesso!')
            return redirect('user_profile')
        else:
            messages.error(request, 'Código inválido. Tente novamente.')
    
    # Gera novo secret se não existir
    if not request.session.get('mfa_temp_secret'):
        secret = pyotp.random_base32()
        request.session['mfa_temp_secret'] = secret
    else:
        secret = request.session['mfa_temp_secret']
    
    # Gera QR code
    totp = pyotp.TOTP(secret)
    issuer_name = "MPLS Search System"
    qr_uri = totp.provisioning_uri(
        name=request.user.email or request.user.username,
        issuer_name=issuer_name
    )
    
    # Cria QR code
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(qr_uri)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    img_str = base64.b64encode(buffer.getvalue()).decode()
    
    context = {
        'qr_code': img_str,
        'secret': secret,
        'manual_entry_key': secret
    }
    
    return render(request, 'mpls_analyzer/setup_mfa.html', context)


@require_mfa
def disable_mfa(request):
    """View para desabilitar MFA"""
    if request.method == 'POST':
        password = request.POST.get('password')
        
        # Verifica senha atual
        if not request.user.check_password(password):
            messages.error(request, 'Senha incorreta.')
            return redirect('user_profile')
        
        profile, created = UserProfile.objects.get_or_create(user=request.user)
        profile.mfa_enabled = False
        profile.require_mfa = False
        profile.mfa_secret = None
        profile.save()
        
        log_security_event(
            'MFA_DISABLED',
            f'MFA desabilitado por {request.user.username}',
            request.user,
            get_client_ip(request)
        )
        
        messages.success(request, 'MFA desabilitado com sucesso!')
        return redirect('user_profile')
    
    return render(request, 'mpls_analyzer/disable_mfa.html')


@require_mfa
def toggle_mfa(request):
    """API para verificar status MFA e redirecionar"""
    if request.method == 'POST':
        profile, created = UserProfile.objects.get_or_create(user=request.user)
        
        if profile.mfa_enabled:
            # Se MFA está habilitado, redireciona para desabilitar
            return JsonResponse({
                'success': True,
                'action': 'redirect',
                'url': '/profile/disable-mfa/',
                'message': 'Redirecionando para desabilitar MFA...'
            })
        else:
            # Se MFA está desabilitado, redireciona para configurar
            return JsonResponse({
                'success': True,
                'action': 'redirect',
                'url': '/profile/setup-mfa/',
                'message': 'Redirecionando para configurar MFA...'
            })
    
    return JsonResponse({'success': False, 'message': 'Método não permitido'})


# =============================================================================
# VIEWS DE LOGS DE ACESSO E AUDITORIA PARA GERENTES
# =============================================================================

def is_manager(user):
    """Verifica se o usuário é gerente/admin"""
    return user.is_authenticated and (user.is_superuser or hasattr(user, 'profile') and user.profile.is_admin)


@require_mfa
@user_passes_test(is_manager)
def audit_dashboard(request):
    """Dashboard dedicado para auditoria e logs"""
    from datetime import datetime, timedelta
    from django.utils import timezone
    
    # Períodos de análise
    last_24h = timezone.now() - timedelta(hours=24)
    last_7d = timezone.now() - timedelta(days=7)
    last_30d = timezone.now() - timedelta(days=30)
    
    # Estatísticas de Access Logs
    access_stats = {
        '24h': {
            'total': AccessLog.objects.filter(login_time__gte=last_24h).count(),
            'success': AccessLog.objects.filter(login_time__gte=last_24h, status='success').count(),
            'failed': AccessLog.objects.filter(login_time__gte=last_24h, status='failed').count(),
            'unique_ips': AccessLog.objects.filter(login_time__gte=last_24h).values('ip_address').distinct().count(),
        },
        '7d': {
            'total': AccessLog.objects.filter(login_time__gte=last_7d).count(),
            'success': AccessLog.objects.filter(login_time__gte=last_7d, status='success').count(),
            'failed': AccessLog.objects.filter(login_time__gte=last_7d, status='failed').count(),
            'unique_ips': AccessLog.objects.filter(login_time__gte=last_7d).values('ip_address').distinct().count(),
        },
        '30d': {
            'total': AccessLog.objects.filter(login_time__gte=last_30d).count(),
            'success': AccessLog.objects.filter(login_time__gte=last_30d, status='success').count(),
            'failed': AccessLog.objects.filter(login_time__gte=last_30d, status='failed').count(),
            'unique_ips': AccessLog.objects.filter(login_time__gte=last_30d).values('ip_address').distinct().count(),
        }
    }
    
    # Estatísticas de Audit Logs
    audit_stats = {
        '24h': {
            'total': AuditLog.objects.filter(timestamp__gte=last_24h).count(),
            'searches': AuditLog.objects.filter(timestamp__gte=last_24h, action='search').count(),
            'exports': AuditLog.objects.filter(timestamp__gte=last_24h, action='report_export').count(),
            'views': AuditLog.objects.filter(timestamp__gte=last_24h, action__in=['view_equipment', 'view_vpn']).count(),
        },
        '7d': {
            'total': AuditLog.objects.filter(timestamp__gte=last_7d).count(),
            'searches': AuditLog.objects.filter(timestamp__gte=last_7d, action='search').count(),
            'exports': AuditLog.objects.filter(timestamp__gte=last_7d, action='report_export').count(),
            'views': AuditLog.objects.filter(timestamp__gte=last_7d, action__in=['view_equipment', 'view_vpn']).count(),
        },
        '30d': {
            'total': AuditLog.objects.filter(timestamp__gte=last_30d).count(),
            'searches': AuditLog.objects.filter(timestamp__gte=last_30d, action='search').count(),
            'exports': AuditLog.objects.filter(timestamp__gte=last_30d, action='report_export').count(),
            'views': AuditLog.objects.filter(timestamp__gte=last_30d, action__in=['view_equipment', 'view_vpn']).count(),
        }
    }
    
    # Logs recentes para o dashboard
    recent_access_logs = AccessLog.objects.select_related('user').order_by('-login_time')[:10]
    recent_audit_logs = AuditLog.objects.select_related('user').order_by('-timestamp')[:10]
    
    # Top IPs por tentativas de login
    top_ips = (AccessLog.objects.filter(login_time__gte=last_7d)
               .values('ip_address')
               .annotate(total=Count('id'), 
                        failed=Count('id', filter=Q(status='failed')))
               .order_by('-total')[:10])
    
    # Usuários mais ativos
    top_users = (AuditLog.objects.filter(timestamp__gte=last_7d)
                .values('user__username')
                .annotate(total_actions=Count('id'),
                         searches=Count('id', filter=Q(action='search')),
                         exports=Count('id', filter=Q(action='report_export')))
                .order_by('-total_actions')[:10])
    
    context = {
        'access_stats': access_stats,
        'audit_stats': audit_stats,
        'recent_access_logs': recent_access_logs,
        'recent_audit_logs': recent_audit_logs,
        'top_ips': top_ips,
        'top_users': top_users,
    }
    
    return render(request, 'mpls_analyzer/audit_dashboard.html', context)


@require_mfa
@user_passes_test(is_manager)
def access_logs_view(request):
    """View para visualizar logs de acesso"""
    # Filtros
    user_filter = request.GET.get('user', '')
    ip_filter = request.GET.get('ip', '')
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Query base
    logs = AccessLog.objects.select_related('user').order_by('-login_time')
    
    # Aplicar filtros
    if user_filter:
        logs = logs.filter(user__username__icontains=user_filter)
    
    if ip_filter:
        logs = logs.filter(ip_address__icontains=ip_filter)
    
    if status_filter:
        logs = logs.filter(status=status_filter)
    
    if date_from:
        from datetime import datetime
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            logs = logs.filter(login_time__date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        from datetime import datetime
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            logs = logs.filter(login_time__date__lte=date_to_obj)
        except ValueError:
            pass
    
    # Paginação
    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estatísticas
    total_logs = logs.count()
    success_logs = logs.filter(status='success').count()
    failed_logs = logs.filter(status='failed').count()
    
    # Usuários únicos para filtro
    users = User.objects.filter(access_logs__isnull=False).distinct().values_list('username', flat=True)
    
    # Registra auditoria da visualização
    AuditLogger.log_user_management(
        user=request.user,
        request=request,
        action_type='view_access_logs',
        target_username=user_filter if user_filter else None
    )
    
    context = {
        'page_obj': page_obj,
        'total_logs': total_logs,
        'success_logs': success_logs,
        'failed_logs': failed_logs,
        'user_filter': user_filter,
        'ip_filter': ip_filter,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
        'users': users,
        'status_choices': AccessLog.STATUS_CHOICES,
    }
    
    return render(request, 'mpls_analyzer/access_logs.html', context)


@require_mfa
@user_passes_test(is_manager)
def audit_logs_view(request):
    """View para visualizar logs de auditoria"""
    # Filtros
    user_filter = request.GET.get('user', '')
    action_filter = request.GET.get('action', '')
    ip_filter = request.GET.get('ip', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_query = request.GET.get('search', '')
    
    # Query base
    logs = AuditLog.objects.select_related('user').order_by('-timestamp')
    
    # Aplicar filtros
    if user_filter:
        logs = logs.filter(user__username__icontains=user_filter)
    
    if action_filter:
        logs = logs.filter(action=action_filter)
    
    if ip_filter:
        logs = logs.filter(ip_address__icontains=ip_filter)
    
    if search_query:
        logs = logs.filter(
            Q(description__icontains=search_query) |
            Q(search_query__icontains=search_query) |
            Q(target_object_type__icontains=search_query)
        )
    
    if date_from:
        from datetime import datetime
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            logs = logs.filter(timestamp__date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        from datetime import datetime
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            logs = logs.filter(timestamp__date__lte=date_to_obj)
        except ValueError:
            pass
    
    # Paginação
    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estatísticas
    total_logs = logs.count()
    search_logs = logs.filter(action='search').count()
    export_logs = logs.filter(action='report_export').count()
    
    # Usuários únicos para filtro
    users = User.objects.filter(audit_logs__isnull=False).distinct().values_list('username', flat=True)
    
    # Registra auditoria da visualização
    AuditLogger.log_user_management(
        user=request.user,
        request=request,
        action_type='view_audit_logs',
        target_username=user_filter if user_filter else None
    )
    
    context = {
        'page_obj': page_obj,
        'total_logs': total_logs,
        'search_logs': search_logs,
        'export_logs': export_logs,
        'user_filter': user_filter,
        'action_filter': action_filter,
        'ip_filter': ip_filter,
        'date_from': date_from,
        'date_to': date_to,
        'search_query': search_query,
        'users': users,
        'action_choices': AuditLog.ACTION_CHOICES,
    }
    
    return render(request, 'mpls_analyzer/audit_logs.html', context)


@require_mfa
@user_passes_test(is_manager)
def export_access_logs(request):
    """Exporta logs de acesso para Excel"""
    # Mesmo filtros da view principal
    user_filter = request.GET.get('user', '')
    ip_filter = request.GET.get('ip', '')
    status_filter = request.GET.get('status', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Query com filtros
    logs = AccessLog.objects.select_related('user').order_by('-login_time')
    
    if user_filter:
        logs = logs.filter(user__username__icontains=user_filter)
    if ip_filter:
        logs = logs.filter(ip_address__icontains=ip_filter)
    if status_filter:
        logs = logs.filter(status=status_filter)
    if date_from:
        from datetime import datetime
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            logs = logs.filter(login_time__date__gte=date_from_obj)
        except ValueError:
            pass
    if date_to:
        from datetime import datetime
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            logs = logs.filter(login_time__date__lte=date_to_obj)
        except ValueError:
            pass
    
    # Cria o Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Logs de Acesso"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Cabeçalho
    headers = ['Usuário', 'IP', 'Data/Hora Login', 'Data/Hora Logout', 'Status', 'Motivo da Falha', 'User Agent']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Dados
    row = 2
    for log in logs[:5000]:  # Limita a 5000 registros
        data_row = [
            log.user.username if log.user else 'N/A',
            log.ip_address,
            log.login_time.strftime('%d/%m/%Y %H:%M:%S'),
            log.logout_time.strftime('%d/%m/%Y %H:%M:%S') if log.logout_time else 'N/A',
            log.get_status_display(),
            log.failure_reason or 'N/A',
            log.user_agent[:100] + '...' if len(log.user_agent) > 100 else log.user_agent
        ]
        
        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = center_alignment
        
        row += 1
    
    # Ajusta largura das colunas
    column_widths = [20, 15, 20, 20, 15, 30, 50]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Registra auditoria da exportação
    AuditLogger.log_report_export(
        user=request.user,
        request=request,
        export_format='xlsx',
        results_count=logs.count(),
        report_type='access_logs'
    )
    
    # Resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="logs_acesso_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@require_mfa
@user_passes_test(is_manager)
def export_audit_logs(request):
    """Exporta logs de auditoria para Excel"""
    # Mesmo filtros da view principal
    user_filter = request.GET.get('user', '')
    action_filter = request.GET.get('action', '')
    ip_filter = request.GET.get('ip', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_query = request.GET.get('search', '')
    
    # Query com filtros
    logs = AuditLog.objects.select_related('user').order_by('-timestamp')
    
    if user_filter:
        logs = logs.filter(user__username__icontains=user_filter)
    if action_filter:
        logs = logs.filter(action=action_filter)
    if ip_filter:
        logs = logs.filter(ip_address__icontains=ip_filter)
    if search_query:
        logs = logs.filter(
            Q(description__icontains=search_query) |
            Q(search_query__icontains=search_query) |
            Q(target_object_type__icontains=search_query)
        )
    if date_from:
        from datetime import datetime
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            logs = logs.filter(timestamp__date__gte=date_from_obj)
        except ValueError:
            pass
    if date_to:
        from datetime import datetime
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            logs = logs.filter(timestamp__date__lte=date_to_obj)
        except ValueError:
            pass
    
    # Cria o Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Logs de Auditoria"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Cabeçalho
    headers = ['Usuário', 'Ação', 'Descrição', 'IP', 'Data/Hora', 'Objeto Alvo', 'Consulta/Busca', 'Formato Export', 'Qtd Resultados']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Dados
    row = 2
    for log in logs[:5000]:  # Limita a 5000 registros
        data_row = [
            log.user.username if log.user else 'N/A',
            log.get_action_display(),
            log.description[:200] + '...' if len(log.description) > 200 else log.description,
            log.ip_address,
            log.timestamp.strftime('%d/%m/%Y %H:%M:%S'),
            f"{log.target_object_type} #{log.target_object_id}" if log.target_object_type and log.target_object_id else 'N/A',
            log.search_query[:100] + '...' if len(log.search_query) > 100 else log.search_query or 'N/A',
            log.export_format.upper() if log.export_format else 'N/A',
            log.results_count or 'N/A'
        ]
        
        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = center_alignment
        
        row += 1
    
    # Ajusta largura das colunas
    column_widths = [15, 20, 30, 15, 20, 20, 30, 15, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Registra auditoria da exportação
    AuditLogger.log_report_export(
        user=request.user,
        request=request,
        export_format='xlsx',
        results_count=logs.count(),
        report_type='audit_logs'
    )
    
    # Resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="logs_auditoria_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response
